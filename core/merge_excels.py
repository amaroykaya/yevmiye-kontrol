import logging
import os
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


OUTPUT_COLUMNS = [
    "Sıra No",
    "Boş",
    "Tarih",
    "Fatura No",
    "Firma",
    "KDV",
    "KDV'siz",
    "Toplam",
    "Etiket",
    "Ödeme Kanalı",
]


def _to_float(value: object) -> float:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return 0.0
    text = text.replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _format_date(value: object) -> str:
    if isinstance(value, (datetime, pd.Timestamp)):
        return value.strftime("%d.%m.%Y")
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return ""
    return text.replace("/", ".")


def _safe_cell(df: pd.DataFrame, row_idx: int, col_idx: int) -> object:
    if col_idx >= df.shape[1]:
        return ""
    return df.iat[row_idx, col_idx]


def read_gider_rows(file_path: str, logger: logging.Logger) -> list[dict]:
    logger.info("Gider dosyası okunuyor: %s", file_path)
    df = pd.read_excel(file_path, dtype=object, header=None)
    rows: list[dict] = []

    # 5. satırdan (1-indexed) başlanır.
    for i in range(4, len(df)):
        try:
            b_val = str(_safe_cell(df, i, 1)).strip()  # B
            if not b_val or b_val.lower() == "nan" or b_val == "-":
                continue

            fatura_no = str(_safe_cell(df, i, 10)).strip()  # K
            if not fatura_no or fatura_no.lower() == "nan":
                continue

            rows.append(
                {
                    "Tarih": _format_date(_safe_cell(df, i, 3)),  # D
                    "Fatura No": fatura_no,
                    "Firma": str(_safe_cell(df, i, 4)).strip(),  # E
                    "KDV": _to_float(_safe_cell(df, i, 11)),  # L
                    "KDV'siz": _to_float(_safe_cell(df, i, 12)),  # M
                    "Toplam": _to_float(_safe_cell(df, i, 8)),  # I
                    "Etiket": str(_safe_cell(df, i, 13)).strip(),  # N
                    "Ödeme Kanalı": str(_safe_cell(df, i, 9)).strip(),  # J
                }
            )
        except Exception as exc:
            logger.warning("Gider satırı atlandı (index=%s): %s", i, exc)

    logger.info("Gider satır sayısı: %s", len(rows))
    return rows


def read_gelir_rows(file_path: str, logger: logging.Logger) -> list[dict]:
    logger.info("Gelir dosyası okunuyor: %s", file_path)
    df = pd.read_excel(file_path, dtype=object, header=None)
    rows: list[dict] = []

    # 2. satırdan (1-indexed) başlanır.
    for i in range(1, len(df)):
        try:
            b_val = str(_safe_cell(df, i, 1)).strip()  # B
            if not b_val or b_val.lower() == "nan" or b_val == "-":
                continue

            fatura_no = str(_safe_cell(df, i, 6)).strip()  # G
            if not fatura_no or fatura_no.lower() == "nan":
                continue

            rows.append(
                {
                    "Tarih": _format_date(_safe_cell(df, i, 2)),  # C
                    "Fatura No": fatura_no,
                    "Firma": str(_safe_cell(df, i, 3)).strip(),  # D
                    "KDV": _to_float(_safe_cell(df, i, 8)),  # I
                    "KDV'siz": _to_float(_safe_cell(df, i, 9)),  # J
                    "Toplam": _to_float(_safe_cell(df, i, 10)),  # K
                    "Etiket": "",
                    "Ödeme Kanalı": "",
                }
            )
        except Exception as exc:
            logger.warning("Gelir satırı atlandı (index=%s): %s", i, exc)

    logger.info("Gelir satır sayısı: %s", len(rows))
    return rows


def build_combined_rows(gider_rows: list[dict], gelir_rows: list[dict]) -> list[dict]:
    combined: list[dict] = []
    ordered_rows = gider_rows + gelir_rows

    for idx, row in enumerate(ordered_rows, start=1):
        combined.append(
            {
                "Sıra No": idx,
                "Boş": "",
                "Tarih": row.get("Tarih", ""),
                "Fatura No": row.get("Fatura No", ""),
                "Firma": row.get("Firma", ""),
                "KDV": float(row.get("KDV", 0.0)),
                "KDV'siz": float(row.get("KDV'siz", 0.0)),
                "Toplam": float(row.get("Toplam", 0.0)),
                "Etiket": row.get("Etiket", ""),
                "Ödeme Kanalı": row.get("Ödeme Kanalı", ""),
            }
        )

    return combined


def write_combined_excel(combined_rows: list[dict], output_dir: str, logger: logging.Logger) -> str:
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "birlestirilmis_kontrol_verisi.xlsx")
    df = pd.DataFrame(combined_rows, columns=OUTPUT_COLUMNS)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Kontrol")
        ws = writer.sheets["Kontrol"]

        for cell in ws[1]:
            cell.font = Font(bold=True)

        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
            values = df[col_name].astype(str).tolist() if not df.empty else []
            max_len = max([len(col_name), *[len(v) for v in values]] if values else [len(col_name)])
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 80)

    logger.info("Birleştirilmiş excel yazıldı: %s", output_path)
    return output_path


def merge_gider_gelir_excels(
    gider_file_path: str, gelir_file_path: str, output_dir: str, logger: logging.Logger
) -> str:
    gider_rows = read_gider_rows(gider_file_path, logger)
    gelir_rows = read_gelir_rows(gelir_file_path, logger)
    combined_rows = build_combined_rows(gider_rows, gelir_rows)
    return write_combined_excel(combined_rows, output_dir, logger)
