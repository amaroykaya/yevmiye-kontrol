import logging
import os
import re
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from core.fis_summary import build_fis_summary_list
from core.merge_excels import build_combined_rows, read_gelir_rows, read_gider_rows
from core.yevmiye_parser import parse_fis_blocks

STATUS_TAM = "TAM_UYUMLU"
STATUS_FARK = "FARK_VAR"
STATUS_YOK = "ESLESME_YOK"

FILL_GREEN = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
FILL_YELLOW = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")
FILL_RED = PatternFill(fill_type="solid", start_color="F8CBAD", end_color="F8CBAD")

RESULT_COLUMNS = [
    "Sıra No",
    "Tarih",
    "Fatura No",
    "Firma",
    "KDV",
    "KDV'siz",
    "Toplam",
    "Etiket",
    "Ödeme Kanalı",
    "Durum",
    "Eşleşen Satır No",
    "Toplam Kontrol",
    "KDV Kontrol",
    "Etiket Kontrol",
    "Not",
]

MONTH_NAMES_TR = {
    1: "ocak",
    2: "şubat",
    3: "mart",
    4: "nisan",
    5: "mayıs",
    6: "haziran",
    7: "temmuz",
    8: "ağustos",
    9: "eylül",
    10: "ekim",
    11: "kasım",
    12: "aralık",
}


def _norm_text(value: object) -> str:
    text = str(value or "").strip()
    return " ".join(text.split()).lower()


def _normalize_fatura_no(value: object) -> str:
    if value is None:
        return None

    s = str(value).upper().strip()

    # özel karakterleri sil
    s = re.sub(r"[\s\-\./,]", "", s)

    if not s:
        return None

    # tamamen sayıysa
    if s.isdigit():
        return str(int(s))  # baştaki sıfırlar gider

    # harf + sayı karışık
    parts = re.findall(r"[A-Z]+|\d+", s)
    normalized_parts = []

    for p in parts:
        if p.isdigit():
            normalized_parts.append(str(int(p)))  # sıfırları temizle
        else:
            normalized_parts.append(p)

    normalized = "".join(normalized_parts)
    return normalized or None


def _normalize_firma(value: object) -> str:
    text = str(value or "").strip().upper()
    return " ".join(text.split())


def _normalize_tr_text(value: str) -> str:
    return (
        value.lower()
        .replace("ş", "s")
        .replace("ğ", "g")
        .replace("ı", "i")
        .replace("ö", "o")
        .replace("ü", "u")
        .replace("ç", "c")
    )


def _detect_month_from_filename(file_path: str) -> str | None:
    filename = os.path.basename(file_path)
    lowered = filename.lower()
    normalized = _normalize_tr_text(lowered)

    month_word_map = {
        "ocak": "ocak",
        "subat": "şubat",
        "mart": "mart",
        "nisan": "nisan",
        "mayis": "mayıs",
        "haziran": "haziran",
        "temmuz": "temmuz",
        "agustos": "ağustos",
        "eylul": "eylül",
        "ekim": "ekim",
        "kasim": "kasım",
        "aralik": "aralık",
    }

    for key, val in month_word_map.items():
        if key in normalized:
            return val

    month_num_map = {
        "01": "ocak",
        "02": "şubat",
        "03": "mart",
        "04": "nisan",
        "05": "mayıs",
        "06": "haziran",
        "07": "temmuz",
        "08": "ağustos",
        "09": "eylül",
        "10": "ekim",
        "11": "kasım",
        "12": "aralık",
    }

    patterns = [
        r"(?:19|20)\d{2}[-_./](0[1-9]|1[0-2])",
        r"(0[1-9]|1[0-2])[-_./](?:19|20)\d{2}",
        r"(?<!\d)(0[1-9]|1[0-2])(?!\d)",
    ]
    for pattern in patterns:
        match = re.search(pattern, normalized)
        if match:
            month_num = match.group(1)
            return month_num_map.get(month_num)

    return None


def _to_float(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return 0.0
    text = text.replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _is_prefix_match(left: str | None, right: str | None) -> bool:
    if not left or not right:
        return False
    return left.startswith(right) or right.startswith(left)


def _status_fill(status: str) -> PatternFill:
    if status == STATUS_TAM:
        return FILL_GREEN
    if status == STATUS_FARK:
        return FILL_YELLOW
    return FILL_RED


def _build_yevmiye_rows(yevmiye_file_path: str, logger: logging.Logger) -> list[dict]:
    fisler = parse_fis_blocks(yevmiye_file_path, logger)
    summaries = build_fis_summary_list(fisler, yevmiye_file_path, logger)
    rows: list[dict] = []
    for item in summaries:
        rows.append(
            {
                "Sıra No": item.get("sira_no", 0),
                "Tarih": item.get("tarih", ""),
                "Fatura No": item.get("fatura_no", ""),
                "Firma": item.get("firma", ""),
                "KDV": _to_float(item.get("kdv", 0)),
                "KDV'siz": _to_float(item.get("mal_hizmet", 0)),
                "Toplam": _to_float(item.get("toplam", 0)),
                "Etiket": item.get("etiket", ""),
                "Ödeme Kanalı": "",
                "_hesap_kodu": str(item.get("hesap_kodu", "")),
            }
        )
    return rows


def _write_yevmiye_ozet_excel(
    yevmiye_rows: list[dict], output_dir: str, ay_ismi: str, logger: logging.Logger
) -> str:
    output_path = os.path.join(output_dir, f"yevmiye_ozet_{ay_ismi}.xlsx")
    columns = [
        "Sıra No",
        "Hesap Kodu",
        "Tarih",
        "Fatura No",
        "Firma",
        "KDV",
        "Mal/Hizmet",
        "Toplam",
        "Etiket",
    ]

    export_rows = []
    for row in yevmiye_rows:
        export_rows.append(
            {
                "Sıra No": row.get("Sıra No", ""),
                "Hesap Kodu": row.get("_hesap_kodu", ""),
                "Tarih": row.get("Tarih", ""),
                "Fatura No": row.get("Fatura No", ""),
                "Firma": row.get("Firma", ""),
                "KDV": _to_float(row.get("KDV", 0)),
                "Mal/Hizmet": _to_float(row.get("KDV'siz", 0)),
                "Toplam": _to_float(row.get("Toplam", 0)),
                "Etiket": row.get("Etiket", ""),
            }
        )

    df = pd.DataFrame(export_rows, columns=columns)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Yevmiye_Ozet")
        ws = writer.sheets["Yevmiye_Ozet"]
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for col_idx, col_name in enumerate(columns, start=1):
            values = df[col_name].astype(str).tolist() if not df.empty else []
            max_len = max([len(col_name), *[len(v) for v in values]] if values else [len(col_name)])
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 80)

    logger.info("Yevmiye özet çıktısı oluşturuldu: %s", output_path)
    return output_path


def _build_muhasebe_rows(gider_file_path: str, gelir_file_path: str, logger: logging.Logger) -> list[dict]:
    gider_rows = read_gider_rows(gider_file_path, logger)
    gelir_rows = read_gelir_rows(gelir_file_path, logger)
    logger.info("Gider satır sayısı: %s", len(gider_rows))
    logger.info("Gelir satır sayısı: %s", len(gelir_rows))
    combined = build_combined_rows(gider_rows, gelir_rows)
    logger.info("Birleşik satır sayısı: %s", len(combined))
    return combined


def _match_rows(
    yevmiye_rows: list[dict], muhasebe_rows: list[dict], logger: logging.Logger
) -> tuple[dict[int, int], set[int]]:
    matches: dict[int, int] = {}
    used_muhasebe_indexes: set[int] = set()

    def pick_best(candidates: list[int], y_row: dict) -> int | None:
        if not candidates:
            return None
        y_toplam = _to_float(y_row.get("Toplam", 0))
        y_firma = _normalize_firma(y_row.get("Firma", ""))

        def firma_score(idx: int) -> int:
            m_firma = _normalize_firma(muhasebe_rows[idx].get("Firma", ""))
            return 1 if y_firma and m_firma and y_firma == m_firma else 0

        return min(
            candidates,
            key=lambda idx: (
                abs(y_toplam - _to_float(muhasebe_rows[idx].get("Toplam", 0))),
                -firma_score(idx),
            ),
        )

    for yi, y_row in enumerate(yevmiye_rows):
        normalized_fatura = _normalize_fatura_no(y_row.get("Fatura No", ""))
        if not normalized_fatura:
            continue
        candidate_indexes = [
            mi
            for mi, m_row in enumerate(muhasebe_rows)
            if mi not in used_muhasebe_indexes
            and _normalize_fatura_no(m_row.get("Fatura No", "")) == normalized_fatura
        ]
        if len(candidate_indexes) == 1:
            best_idx = candidate_indexes[0]
            matches[yi] = best_idx
            used_muhasebe_indexes.add(best_idx)
            logger.info(
                "Eşleşme (fatura no): Yevmiye %s -> Muhasebe %s | normalize fatura no ile tek aday bulundu, direkt eşleştirildi | fatura_no=%s",
                yi + 1,
                best_idx + 1,
                normalized_fatura,
            )
            continue

        if len(candidate_indexes) > 1:
            best_idx = pick_best(candidate_indexes, y_row)
            if best_idx is not None:
                matches[yi] = best_idx
                used_muhasebe_indexes.add(best_idx)
                y_toplam = _to_float(y_row.get("Toplam", 0))
                chosen_diff = abs(y_toplam - _to_float(muhasebe_rows[best_idx].get("Toplam", 0)))
                y_firma = _normalize_firma(y_row.get("Firma", ""))
                m_firma = _normalize_firma(muhasebe_rows[best_idx].get("Firma", ""))
                if y_firma and m_firma and y_firma == m_firma:
                    logger.info(
                        "Eşleşme (fatura no): Yevmiye %s -> Muhasebe %s | birden fazla aday bulundu, firma benzerliği ile karar verildi | toplam_fark=%.2f",
                        yi + 1,
                        best_idx + 1,
                        chosen_diff,
                    )
                else:
                    logger.info(
                        "Eşleşme (fatura no): Yevmiye %s -> Muhasebe %s | birden fazla aday bulundu, toplam farkı en küçük olan seçildi | toplam_fark=%.2f",
                        yi + 1,
                        best_idx + 1,
                        chosen_diff,
                    )
            continue

        # Soft match: normalize sonrası birebir yoksa prefix + toplam toleransı ile eşleştir.
        soft_candidates = []
        y_toplam = _to_float(y_row.get("Toplam", 0))
        for mi, m_row in enumerate(muhasebe_rows):
            if mi in used_muhasebe_indexes:
                continue
            m_norm_fatura = _normalize_fatura_no(m_row.get("Fatura No", ""))
            if not _is_prefix_match(normalized_fatura, m_norm_fatura):
                continue
            m_toplam = _to_float(m_row.get("Toplam", 0))
            if abs(y_toplam - m_toplam) <= 1:
                soft_candidates.append(mi)

        if len(soft_candidates) == 1:
            best_idx = soft_candidates[0]
            matches[yi] = best_idx
            used_muhasebe_indexes.add(best_idx)
            logger.info(
                "Eşleşme (soft match): Yevmiye %s -> Muhasebe %s | soft match ile eşleşti",
                yi + 1,
                best_idx + 1,
            )
            continue

        if len(soft_candidates) > 1:
            best_idx = pick_best(soft_candidates, y_row)
            if best_idx is not None:
                matches[yi] = best_idx
                used_muhasebe_indexes.add(best_idx)
                logger.info(
                    "Eşleşme (soft match): Yevmiye %s -> Muhasebe %s | soft match ile birden fazla aday içinden seçildi",
                    yi + 1,
                    best_idx + 1,
                )
            continue

    for yi, y_row in enumerate(yevmiye_rows):
        if yi in matches:
            continue
        y_firma = _norm_text(y_row.get("Firma", ""))
        y_toplam = _to_float(y_row.get("Toplam", 0))
        if not y_firma:
            continue

        fallback_candidates = []
        for mi, m_row in enumerate(muhasebe_rows):
            if mi in used_muhasebe_indexes:
                continue
            if _norm_text(m_row.get("Firma", "")) != y_firma:
                continue
            m_toplam = _to_float(m_row.get("Toplam", 0))
            if abs(y_toplam - m_toplam) <= 1:
                fallback_candidates.append(mi)

        best_idx = pick_best(fallback_candidates, y_row)
        if best_idx is not None:
            matches[yi] = best_idx
            used_muhasebe_indexes.add(best_idx)
            logger.info(
                "Eşleşme (yedek toplam+firma): Yevmiye %s -> Muhasebe %s | firma=%s",
                yi + 1,
                best_idx + 1,
                y_row.get("Firma", ""),
            )

    return matches, used_muhasebe_indexes


def _last_chance_match(
    yevmiye_rows: list[dict],
    muhasebe_rows: list[dict],
    matches: dict[int, int],
    used_muhasebe_indexes: set[int],
    logger: logging.Logger,
) -> dict[int, int]:
    """
    Son şans eşleşme:
    - Sadece normal akışta eşleşmeyen satırlar arasında çalışır
    - Toplam farkı ±1
    - Normalize fatura no son 3 veya 4 hanesi aynı
    """
    extra_matches: dict[int, int] = {}

    unmatched_yevmiye = [yi for yi in range(len(yevmiye_rows)) if yi not in matches]
    unmatched_muhasebe = [mi for mi in range(len(muhasebe_rows)) if mi not in used_muhasebe_indexes]

    def tail_match(a: str | None, b: str | None) -> bool:
        if not a or not b:
            return False
        if len(a) >= 4 and len(b) >= 4 and a[-4:] == b[-4:]:
            return True
        if len(a) >= 3 and len(b) >= 3 and a[-3:] == b[-3:]:
            return True
        return False

    for yi in unmatched_yevmiye:
        y_row = yevmiye_rows[yi]
        y_fatura = _normalize_fatura_no(y_row.get("Fatura No", ""))
        y_toplam = _to_float(y_row.get("Toplam", 0))
        if not y_fatura:
            continue

        candidates: list[int] = []

        for mi in unmatched_muhasebe:
            if mi in used_muhasebe_indexes:
                continue
            m_row = muhasebe_rows[mi]
            m_fatura = _normalize_fatura_no(m_row.get("Fatura No", ""))
            m_toplam = _to_float(m_row.get("Toplam", 0))
            if not tail_match(y_fatura, m_fatura):
                continue
            if abs(y_toplam - m_toplam) <= 1:
                candidates.append(mi)

        if not candidates:
            continue

        best_mi = min(candidates, key=lambda mi: abs(y_toplam - _to_float(muhasebe_rows[mi].get("Toplam", 0))))
        extra_matches[yi] = best_mi
        used_muhasebe_indexes.add(best_mi)
        logger.info("son şans eşleşme: Yevmiye %s -> Muhasebe %s", yi + 1, best_mi + 1)

    return extra_matches


def _build_result_row(base_row: dict, status: str, matched_no: str, toplam: str, kdv: str, etiket: str, note: str) -> dict:
    return {
        "Sıra No": base_row.get("Sıra No", ""),
        "Tarih": base_row.get("Tarih", ""),
        "Fatura No": base_row.get("Fatura No", ""),
        "Firma": base_row.get("Firma", ""),
        "KDV": _to_float(base_row.get("KDV", 0)),
        "KDV'siz": _to_float(base_row.get("KDV'siz", 0)),
        "Toplam": _to_float(base_row.get("Toplam", 0)),
        "Etiket": base_row.get("Etiket", ""),
        "Ödeme Kanalı": base_row.get("Ödeme Kanalı", ""),
        "Durum": status,
        "Eşleşen Satır No": matched_no,
        "Toplam Kontrol": toplam,
        "KDV Kontrol": kdv,
        "Etiket Kontrol": etiket,
        "Not": note,
        "_hesap_kodu": str(base_row.get("_hesap_kodu", "")),
        "_fatura_ok": base_row.get("_fatura_ok", False),
        "_firma_ok": base_row.get("_firma_ok", False),
    }


def _analyze_unmatched_reason(
    source: str, row_no: int, row: dict, counterpart_rows: list[dict], logger: logging.Logger
) -> str:
    raw_fatura = str(row.get("Fatura No", "")).strip()
    norm_fatura = _normalize_fatura_no(raw_fatura)
    raw_firma = str(row.get("Firma", "")).strip()
    norm_firma = _normalize_firma(raw_firma)
    toplam = _to_float(row.get("Toplam", 0))
    kdv = _to_float(row.get("KDV", 0))

    logger.info(
        "ESLESME_YOK DEBUG | kaynak=%s | satır=%s | fatura_no_ham=%s | fatura_no_norm=%s | firma_ham=%s | firma_norm=%s | toplam=%.2f | kdv=%.2f",
        source,
        row_no,
        raw_fatura,
        norm_fatura,
        raw_firma,
        norm_firma,
        toplam,
        kdv,
    )

    same_fatura_candidates = []
    if norm_fatura:
        same_fatura_candidates = [
            idx
            for idx, c_row in enumerate(counterpart_rows)
            if _normalize_fatura_no(c_row.get("Fatura No", "")) == norm_fatura
        ]
    logger.info(
        "ESLESME_YOK DEBUG | kaynak=%s | satır=%s | aynı_fatura_no_aday_sayısı=%s",
        source,
        row_no,
        len(same_fatura_candidates),
    )

    if len(same_fatura_candidates) > 1:
        return "birden fazla aday var"
    if len(same_fatura_candidates) == 1:
        return "normalize sonrası da eşleşme yok"

    total_near_candidates = []
    for idx, c_row in enumerate(counterpart_rows):
        c_toplam = _to_float(c_row.get("Toplam", 0))
        if abs(toplam - c_toplam) <= 1:
            total_near_candidates.append(idx)

    if total_near_candidates:
        logger.info(
            "ESLESME_YOK DEBUG | kaynak=%s | satır=%s | toplam_±1_aday_sayısı=%s",
            source,
            row_no,
            len(total_near_candidates),
        )
        same_firma_near = [
            idx
            for idx in total_near_candidates
            if _normalize_firma(counterpart_rows[idx].get("Firma", "")) == norm_firma and norm_firma
        ]
        if same_firma_near:
            logger.info(
                "ESLESME_YOK DEBUG | kaynak=%s | satır=%s | toplam_yakın_ve_firma_benzer_aday_var",
                source,
                row_no,
            )
            return "normalize sonrası da eşleşme yok"

        logger.info(
            "ESLESME_YOK DEBUG | kaynak=%s | satır=%s | toplam yakın ama firma farklı",
            source,
            row_no,
        )
        return "toplam yakın aday var"

    if norm_firma:
        firma_similar_but_fatura_diff = [
            idx
            for idx, c_row in enumerate(counterpart_rows)
            if _normalize_firma(c_row.get("Firma", "")) == norm_firma
            and _normalize_fatura_no(c_row.get("Fatura No", "")) != norm_fatura
        ]
        if firma_similar_but_fatura_diff:
            logger.info(
                "ESLESME_YOK DEBUG | kaynak=%s | satır=%s | firma benzer ama fatura_no farklı",
                source,
                row_no,
            )
            return "firma farklı"

    logger.info(
        "ESLESME_YOK DEBUG | kaynak=%s | satır=%s | normalize sonrası da eşleşme yok",
        source,
        row_no,
    )
    return "fatura no bulunamadı"


def _compare(
    yevmiye_rows: list[dict],
    muhasebe_rows: list[dict],
    matches: dict[int, int],
    last_chance_match_keys: set[int],
    logger: logging.Logger,
) -> tuple[list[dict], list[dict]]:
    yevmiye_result: list[dict] = []
    muhasebe_result: list[dict] = []

    matched_muhasebe = {mi: yi for yi, mi in matches.items()}
    tam_count = 0
    fark_count = 0
    yok_count = 0

    for yi, y_row in enumerate(yevmiye_rows):
        if yi not in matches:
            yok_count += 1
            note = _analyze_unmatched_reason("Yevmiye", yi + 1, y_row, muhasebe_rows, logger)
            logger.info("Eşleşme yok (kırmızı): Yevmiye %s | fatura_no=%s", yi + 1, y_row.get("Fatura No", ""))
            yevmiye_result.append(
                _build_result_row(y_row, STATUS_YOK, "", "-", "-", "-", note)
            )
            continue

        mi = matches[yi]
        m_row = muhasebe_rows[mi]
        toplam_fark = abs(_to_float(y_row.get("Toplam", 0)) - _to_float(m_row.get("Toplam", 0)))
        kdv_fark = abs(_to_float(y_row.get("KDV", 0)) - _to_float(m_row.get("KDV", 0)))

        toplam_ok = toplam_fark <= 1
        kdv_ok = kdv_fark <= 1
        y_kdv = _to_float(y_row.get("KDV", 0))
        m_kdv = _to_float(m_row.get("KDV", 0))
        kdv_special = m_kdv == 0 and y_kdv > 0 and toplam_ok

        y_etiket = str(y_row.get("Etiket", "")).strip()
        m_etiket = str(m_row.get("Etiket", "")).strip()
        etiket_ok = (not y_etiket and not m_etiket) or (y_etiket == m_etiket)
        fatura_ok = _normalize_fatura_no(y_row.get("Fatura No", "")) == _normalize_fatura_no(m_row.get("Fatura No", ""))
        firma_ok = _normalize_firma(y_row.get("Firma", "")) == _normalize_firma(m_row.get("Firma", ""))

        if yi in last_chance_match_keys:
            status = STATUS_FARK
            note = "son hane eşleşmesi ile bulundu"
            fark_count += 1
        elif toplam_ok and kdv_ok and etiket_ok:
            status = STATUS_TAM
            note = "Tüm kontroller uyumlu"
            tam_count += 1
        else:
            status = STATUS_FARK
            reasons = []
            if not toplam_ok:
                reasons.append("toplam farkı")
            if kdv_special:
                reasons.append("yevmiye tarafında kısmi/ek KDV var")
            elif not kdv_ok:
                reasons.append("KDV farkı")
            if not etiket_ok:
                reasons.append("etiket farkı")
            note = ", ".join(reasons)
            fark_count += 1
            logger.info(
                "Eşleşme farkı (sarı): Yevmiye %s -> Muhasebe %s | neden=%s",
                yi + 1,
                mi + 1,
                note,
            )

        yevmiye_result.append(
            _build_result_row(
                {**y_row, "_fatura_ok": fatura_ok, "_firma_ok": firma_ok},
                status,
                str(mi + 1),
                "UYUMLU" if toplam_ok else f"FARK ({toplam_fark:.2f})",
                "KISMI_KDV" if kdv_special else ("UYUMLU" if kdv_ok else f"FARK ({kdv_fark:.2f})"),
                "UYUMLU" if etiket_ok else "FARK",
                note,
            )
        )

    for mi, m_row in enumerate(muhasebe_rows):
        if mi not in matched_muhasebe:
            note = _analyze_unmatched_reason("Muhasebe", mi + 1, m_row, yevmiye_rows, logger)
            muhasebe_result.append(
                _build_result_row(m_row, STATUS_YOK, "", "-", "-", "-", note)
            )
            continue

        yi = matched_muhasebe[mi]
        y_row = yevmiye_rows[yi]
        toplam_fark = abs(_to_float(y_row.get("Toplam", 0)) - _to_float(m_row.get("Toplam", 0)))
        kdv_fark = abs(_to_float(y_row.get("KDV", 0)) - _to_float(m_row.get("KDV", 0)))
        toplam_ok = toplam_fark <= 1
        kdv_ok = kdv_fark <= 1
        y_kdv = _to_float(y_row.get("KDV", 0))
        m_kdv = _to_float(m_row.get("KDV", 0))
        kdv_special = m_kdv == 0 and y_kdv > 0 and toplam_ok
        y_etiket = str(y_row.get("Etiket", "")).strip()
        m_etiket = str(m_row.get("Etiket", "")).strip()
        etiket_ok = (not y_etiket and not m_etiket) or (y_etiket == m_etiket)
        fatura_ok = _normalize_fatura_no(y_row.get("Fatura No", "")) == _normalize_fatura_no(m_row.get("Fatura No", ""))
        firma_ok = _normalize_firma(y_row.get("Firma", "")) == _normalize_firma(m_row.get("Firma", ""))
        if yi in last_chance_match_keys:
            status = STATUS_FARK
            note = "son hane eşleşmesi ile bulundu"
        else:
            status = STATUS_TAM if (toplam_ok and kdv_ok and etiket_ok) else STATUS_FARK
        if status == STATUS_TAM:
            note = "Tüm kontroller uyumlu"
        elif kdv_special:
            note = "yevmiye tarafında kısmi/ek KDV var"
        elif yi not in last_chance_match_keys:
            note = "Kontrol farkı var"
        muhasebe_result.append(
            _build_result_row(
                {**m_row, "_fatura_ok": fatura_ok, "_firma_ok": firma_ok},
                status,
                str(yi + 1),
                "UYUMLU" if toplam_ok else f"FARK ({toplam_fark:.2f})",
                "KISMI_KDV" if kdv_special else ("UYUMLU" if kdv_ok else f"FARK ({kdv_fark:.2f})"),
                "UYUMLU" if etiket_ok else "FARK",
                note,
            )
        )

    logger.info("Kaç tam uyum var: %s", tam_count)
    logger.info("Kaç fark var: %s", fark_count)
    logger.info("Kaç eşleşme yok: %s", yok_count)
    return yevmiye_result, muhasebe_result


def _sort_yevmiye_rows_for_sheet(rows: list[dict]) -> list[dict]:
    # 120 ile başlayan hesap kodları en altta olacak, diğerleri mevcut sırayı koruyacak.
    indexed_rows = list(enumerate(rows))
    indexed_rows.sort(
        key=lambda item: (
            1 if str(item[1].get("_hesap_kodu", "")).strip().startswith("120") else 0,
            item[0],
        )
    )
    sorted_rows = [dict(item[1]) for item in indexed_rows]
    # Sadece sheet çıktısı için sıra numarası sıralama sonrası 1..N olarak yenilenir.
    for idx, row in enumerate(sorted_rows, start=1):
        row["Sıra No"] = idx
    return sorted_rows


def _write_sheet(writer: pd.ExcelWriter, sheet_name: str, rows: list[dict], logger: logging.Logger) -> None:
    df = pd.DataFrame(rows, columns=RESULT_COLUMNS)
    df.to_excel(writer, index=False, sheet_name=sheet_name)
    ws = writer.sheets[sheet_name]

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row_idx in range(2, ws.max_row + 1):
        status = ws.cell(row=row_idx, column=10).value
        fill = _status_fill(str(status))
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill

        # Eşleşen satırlarda (TAM_UYUMLU/FARK_VAR) doğru alanları hücre bazında yeşil göster.
        if str(status) in {STATUS_TAM, STATUS_FARK}:
            toplam_kontrol = str(ws.cell(row=row_idx, column=12).value or "").strip().upper()
            kdv_kontrol = str(ws.cell(row=row_idx, column=13).value or "").strip().upper()
            etiket_kontrol = str(ws.cell(row=row_idx, column=14).value or "").strip().upper()
            eslesen_no = ws.cell(row=row_idx, column=11).value
            fatura_ok = False
            firma_ok = False
            kdv_ok = kdv_kontrol.startswith("OK") or kdv_kontrol.startswith("UYUMLU")
            toplam_ok = toplam_kontrol.startswith("OK") or toplam_kontrol.startswith("UYUMLU")
            etiket_ok = etiket_kontrol.startswith("OK") or etiket_kontrol.startswith("UYUMLU")

            if eslesen_no and row_idx - 2 < len(rows):
                row_meta = rows[row_idx - 2]
                fatura_ok = bool(row_meta.get("_fatura_ok", False))
                firma_ok = bool(row_meta.get("_firma_ok", False))

            # Doğru sütun indeksleri:
            # 1=Sıra No, 2=Tarih, 3=Fatura No, 4=Firma, 5=KDV, 6=KDV'siz, 7=Toplam, 8=Etiket
            if fatura_ok:
                ws.cell(row=row_idx, column=3).fill = FILL_GREEN  # Fatura No
            if firma_ok:
                ws.cell(row=row_idx, column=4).fill = FILL_GREEN  # Firma
            if kdv_ok:
                ws.cell(row=row_idx, column=5).fill = FILL_GREEN  # KDV
            if toplam_ok:
                ws.cell(row=row_idx, column=7).fill = FILL_GREEN  # Toplam
            if etiket_ok:
                ws.cell(row=row_idx, column=8).fill = FILL_GREEN  # Etiket

            logger.info(
                "Hücre boyama: satır %s | fatura=%s firma=%s kdv=%s toplam=%s etiket=%s",
                row_idx - 1,
                fatura_ok,
                firma_ok,
                kdv_ok,
                toplam_ok,
                etiket_ok,
            )

    for col_idx, col_name in enumerate(RESULT_COLUMNS, start=1):
        values = df[col_name].astype(str).tolist() if not df.empty else []
        max_len = max([len(col_name), *[len(v) for v in values]] if values else [len(col_name)])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 80)


def run_reconciliation(
    yevmiye_file_path: str,
    gider_file_path: str,
    gelir_file_path: str,
    output_dir: str,
    logger: logging.Logger,
) -> str:
    logger.info("Seçilen yevmiye dosyası: %s", yevmiye_file_path)
    logger.info("Seçilen gider dosyası: %s", gider_file_path)
    logger.info("Seçilen gelir dosyası: %s", gelir_file_path)
    logger.info("Seçilen çıktı klasörü: %s", output_dir)

    muhasebe_rows = _build_muhasebe_rows(gider_file_path, gelir_file_path, logger)
    yevmiye_rows = _build_yevmiye_rows(yevmiye_file_path, logger)
    logger.info("Yevmiye satır sayısı: %s", len(yevmiye_rows))

    matches, used_muhasebe_indexes = _match_rows(yevmiye_rows, muhasebe_rows, logger)
    last_chance_matches = _last_chance_match(
        yevmiye_rows=yevmiye_rows,
        muhasebe_rows=muhasebe_rows,
        matches=matches,
        used_muhasebe_indexes=used_muhasebe_indexes,
        logger=logger,
    )
    matches.update(last_chance_matches)
    yevmiye_result, muhasebe_result = _compare(
        yevmiye_rows, muhasebe_rows, matches, set(last_chance_matches.keys()), logger
    )
    yevmiye_result_sorted = _sort_yevmiye_rows_for_sheet(yevmiye_result)

    logger.info("Birleşmiş excel toplam satır sayısı: %s", len(muhasebe_rows))
    logger.info("2. sekmeye yazılan satır sayısı: %s", len(muhasebe_result))
    logger.info("Yevmiye özet toplam satır sayısı: %s", len(yevmiye_rows))
    logger.info("1. sekmeye yazılan satır sayısı: %s", len(yevmiye_result_sorted))

    if len(muhasebe_rows) != len(muhasebe_result):
        logger.warning(
            "UYARI: Birleşik satır sayısı ile 2. sekme satır sayısı farklı! kaynak=%s sheet=%s",
            len(muhasebe_rows),
            len(muhasebe_result),
        )
    if len(yevmiye_rows) != len(yevmiye_result_sorted):
        logger.warning(
            "UYARI: Yevmiye satır sayısı ile 1. sekme satır sayısı farklı! kaynak=%s sheet=%s",
            len(yevmiye_rows),
            len(yevmiye_result_sorted),
        )

    os.makedirs(output_dir, exist_ok=True)
    ay_ismi = _detect_month_from_filename(yevmiye_file_path)
    ay_kaynagi = "yevmiye"
    if not ay_ismi:
        ay_ismi = _detect_month_from_filename(gider_file_path)
        ay_kaynagi = "gider"
    if not ay_ismi:
        ay_ismi = _detect_month_from_filename(gelir_file_path)
        ay_kaynagi = "gelir"
    if not ay_ismi:
        ay_ismi = MONTH_NAMES_TR.get(datetime.now().month, "ay")
        ay_kaynagi = "tarih"

    logger.info("Ay tespiti: %s kaynağından alındı (%s)", ay_kaynagi, ay_ismi)
    output_path = os.path.join(output_dir, f"yevmiye_kontrol_{ay_ismi}.xlsx")
    _write_yevmiye_ozet_excel(yevmiye_rows, output_dir, ay_ismi, logger)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        _write_sheet(writer, "Yevmiye_Kontrol", yevmiye_result_sorted, logger)
        _write_sheet(writer, "Birlestirilmis_Excel", muhasebe_result, logger)

    logger.info("Kontrol çıktısı oluşturuldu: %s", output_path)
    return output_path
